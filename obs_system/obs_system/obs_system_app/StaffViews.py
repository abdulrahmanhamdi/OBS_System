import json
from datetime import datetime
from uuid import uuid4
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core import serializers
from django.forms import model_to_dict
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt

from obs_system_app.models import Subjects, SessionYearModel, Students, Attendance, AttendanceReport, \
    LeaveReportStaff, Staffs, FeedBackStaffs, CustomUser, Courses, NotificationStaffs, OnlineClassRoom, StudentSubjectSelection, \
    GradeComponent, StudentResult, StudentComponentResult, NotificationStudent  


def staff_home(request):
    #For Fetch All Student Under Staff
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    course_id_list=[]
    for subject in subjects:
        course=Courses.objects.get(id=subject.course_id.id)
        course_id_list.append(course.id)

    final_course=[]
    #removing Duplicate Course ID
    for course_id in course_id_list:
        if course_id not in final_course:
            final_course.append(course_id)

    students_count=Students.objects.filter(course_id__in=final_course).count()

    #Fetch All Attendance Count
    attendance_count=Attendance.objects.filter(subject_id__in=subjects).count()

    #Fetch All Approve Leave
    staff=Staffs.objects.get(admin=request.user.id)
    leave_count=LeaveReportStaff.objects.filter(staff_id=staff.id,leave_status=1).count()
    subject_count=subjects.count()

    #Fetch Attendance Data by Subject
    subject_list=[]
    attendance_list=[]
    for subject in subjects:
        attendance_count1=Attendance.objects.filter(subject_id=subject.id).count()
        subject_list.append(subject.subject_name)
        attendance_list.append(attendance_count1)

    students_attendance=Students.objects.filter(course_id__in=final_course)
    student_list=[]
    student_list_attendance_present=[]
    student_list_attendance_absent=[]
    for student in students_attendance:
        attendance_present_count=AttendanceReport.objects.filter(status=True,student_id=student.id).count()
        attendance_absent_count=AttendanceReport.objects.filter(status=False,student_id=student.id).count()
        student_list.append(student.admin.username)
        student_list_attendance_present.append(attendance_present_count)
        student_list_attendance_absent.append(attendance_absent_count)

    return render(request,"staff_template/staff_home_template.html",{"students_count":students_count,"attendance_count":attendance_count,"leave_count":leave_count,"subject_count":subject_count,"subject_list":subject_list,"attendance_list":attendance_list,"student_list":student_list,"present_list":student_list_attendance_present,"absent_list":student_list_attendance_absent})

def staff_take_attendance(request):
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    session_years=SessionYearModel.object.all()
    return render(request,"staff_template/staff_take_attendance.html",{"subjects":subjects,"session_years":session_years})

@csrf_exempt
def get_students(request):
    subject_id=request.POST.get("subject")
    session_year=request.POST.get("session_year")

    subject=Subjects.objects.get(id=subject_id)
    session_model=SessionYearModel.object.get(id=session_year)
    students=Students.objects.filter(course_id=subject.course_id,session_year_id=session_model)
    list_data=[]

    for student in students:
        data_small={"id":student.admin.id,"name":student.admin.first_name+" "+student.admin.last_name}
        list_data.append(data_small)
    return JsonResponse(json.dumps(list_data),content_type="application/json",safe=False)

@csrf_exempt
def save_attendance_data(request):
    student_ids=request.POST.get("student_ids")
    subject_id=request.POST.get("subject_id")
    attendance_date=request.POST.get("attendance_date")
    session_year_id=request.POST.get("session_year_id")

    subject_model=Subjects.objects.get(id=subject_id)
    session_model=SessionYearModel.object.get(id=session_year_id)
    json_sstudent=json.loads(student_ids)
    #print(data[0]['id'])


    try:
        attendance=Attendance(subject_id=subject_model,attendance_date=attendance_date,session_year_id=session_model)
        attendance.save()

        for stud in json_sstudent:
             student=Students.objects.get(admin=stud['id'])
             attendance_report=AttendanceReport(student_id=student,attendance_id=attendance,status=stud['status'])
             attendance_report.save()
        return HttpResponse("OK")
    except:
        return HttpResponse("ERR")

def staff_update_attendance(request):
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    session_year_id=SessionYearModel.object.all()
    return render(request,"staff_template/staff_update_attendance.html",{"subjects":subjects,"session_year_id":session_year_id})

@csrf_exempt
def get_attendance_dates(request):
    subject=request.POST.get("subject")
    session_year_id=request.POST.get("session_year_id")
    subject_obj=Subjects.objects.get(id=subject)
    session_year_obj=SessionYearModel.object.get(id=session_year_id)
    attendance=Attendance.objects.filter(subject_id=subject_obj,session_year_id=session_year_obj)
    attendance_obj=[]
    for attendance_single in attendance:
        data={"id":attendance_single.id,"attendance_date":str(attendance_single.attendance_date),"session_year_id":attendance_single.session_year_id.id}
        attendance_obj.append(data)

    return JsonResponse(json.dumps(attendance_obj),safe=False)

@csrf_exempt
def get_attendance_student(request):
    attendance_date=request.POST.get("attendance_date")
    attendance=Attendance.objects.get(id=attendance_date)

    attendance_data=AttendanceReport.objects.filter(attendance_id=attendance)
    list_data=[]

    for student in attendance_data:
        data_small={"id":student.student_id.admin.id,"name":student.student_id.admin.first_name+" "+student.student_id.admin.last_name,"status":student.status}
        list_data.append(data_small)
    return JsonResponse(json.dumps(list_data),content_type="application/json",safe=False)

@csrf_exempt
def save_updateattendance_data(request):
    student_ids=request.POST.get("student_ids")
    attendance_date=request.POST.get("attendance_date")
    attendance=Attendance.objects.get(id=attendance_date)

    json_sstudent=json.loads(student_ids)


    try:
        for stud in json_sstudent:
             student=Students.objects.get(admin=stud['id'])
             attendance_report=AttendanceReport.objects.get(student_id=student,attendance_id=attendance)
             attendance_report.status=stud['status']
             attendance_report.save()
        return HttpResponse("OK")
    except:
        return HttpResponse("ERR")

def staff_apply_leave(request):
    staff_obj = Staffs.objects.get(admin=request.user.id)
    leave_data=LeaveReportStaff.objects.filter(staff_id=staff_obj)
    return render(request,"staff_template/staff_apply_leave.html",{"leave_data":leave_data})

def staff_apply_leave_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("staff_apply_leave"))
    else:
        leave_date=request.POST.get("leave_date")
        leave_msg=request.POST.get("leave_msg")

        staff_obj=Staffs.objects.get(admin=request.user.id)
        try:
            leave_report=LeaveReportStaff(staff_id=staff_obj,leave_date=leave_date,leave_message=leave_msg,leave_status=0)
            leave_report.save()
            messages.success(request, "Successfully Applied for Leave")
            return HttpResponseRedirect(reverse("staff_apply_leave"))
        except:
            messages.error(request, "Failed To Apply for Leave")
            return HttpResponseRedirect(reverse("staff_apply_leave"))


def staff_feedback(request):
    staff_id=Staffs.objects.get(admin=request.user.id)
    feedback_data=FeedBackStaffs.objects.filter(staff_id=staff_id)
    return render(request,"staff_template/staff_feedback.html",{"feedback_data":feedback_data})

def staff_feedback_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("staff_feedback_save"))
    else:
        feedback_msg=request.POST.get("feedback_msg")

        staff_obj=Staffs.objects.get(admin=request.user.id)
        try:
            feedback=FeedBackStaffs(staff_id=staff_obj,feedback=feedback_msg,feedback_reply="")
            feedback.save()
            messages.success(request, "Successfully Sent Feedback")
            return HttpResponseRedirect(reverse("staff_feedback"))
        except:
            messages.error(request, "Failed To Send Feedback")
            return HttpResponseRedirect(reverse("staff_feedback"))

def staff_profile(request):
    user=CustomUser.objects.get(id=request.user.id)
    staff=Staffs.objects.get(admin=user)
    return render(request,"staff_template/staff_profile.html",{"user":user,"staff":staff})

def staff_profile_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("staff_profile"))
    else:
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        address=request.POST.get("address")
        password=request.POST.get("password")
        try:
            customuser=CustomUser.objects.get(id=request.user.id)
            customuser.first_name=first_name
            customuser.last_name=last_name
            if password!=None and password!="":
                customuser.set_password(password)
            customuser.save()

            staff=Staffs.objects.get(admin=customuser.id)
            staff.address=address
            staff.save()
            messages.success(request, "Successfully Updated Profile")
            return HttpResponseRedirect(reverse("staff_profile"))
        except:
            messages.error(request, "Failed to Update Profile")
            return HttpResponseRedirect(reverse("staff_profile"))

@csrf_exempt
def staff_fcmtoken_save(request):
    token=request.POST.get("token")
    try:
        staff=Staffs.objects.get(admin=request.user.id)
        staff.fcm_token=token
        staff.save()
        return HttpResponse("True")
    except:
        return HttpResponse("False")

def staff_all_notification(request):
    staff=Staffs.objects.get(admin=request.user.id)
    notifications=NotificationStaffs.objects.filter(staff_id=staff.id)
    return render(request,"staff_template/all_notification.html",{"notifications":notifications})

def staff_add_result(request):
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    session_years=SessionYearModel.object.all()
    return render(request,"staff_template/staff_add_result.html",{"subjects":subjects,"session_years":session_years})

def save_student_result(request):
    if request.method!='POST':
        return HttpResponseRedirect('staff_add_result')
    student_admin_id=request.POST.get('student_list')
    assignment_marks=request.POST.get('assignment_marks')
    exam_marks=request.POST.get('exam_marks')
    subject_id=request.POST.get('subject')


    student_obj=Students.objects.get(admin=student_admin_id)
    subject_obj=Subjects.objects.get(id=subject_id)

    try:
        check_exist=StudentResult.objects.filter(subject_id=subject_obj,student_id=student_obj).exists()
        if check_exist:
            result=StudentResult.objects.get(subject_id=subject_obj,student_id=student_obj)
            result.subject_assignment_marks=assignment_marks
            result.subject_exam_marks=exam_marks
            result.save()
            messages.success(request, "Successfully Updated Result")
            return HttpResponseRedirect(reverse("staff_add_result"))
        else:
            result=StudentResult(student_id=student_obj,subject_id=subject_obj,subject_exam_marks=exam_marks,subject_assignment_marks=assignment_marks)
            result.save()
            messages.success(request, "Successfully Added Result")
            return HttpResponseRedirect(reverse("staff_add_result"))
    except:
        messages.error(request, "Failed to Add Result")
        return HttpResponseRedirect(reverse("staff_add_result"))

@csrf_exempt
def fetch_result_student(request):
    subject_id=request.POST.get('subject_id')
    student_id=request.POST.get('student_id')
    student_obj=Students.objects.get(admin=student_id)
    result=StudentResult.objects.filter(student_id=student_obj.id,subject_id=subject_id).exists()
    if result:
        result=StudentResult.objects.get(student_id=student_obj.id,subject_id=subject_id)
        result_data={"exam_marks":result.subject_exam_marks,"assign_marks":result.subject_assignment_marks}
        return HttpResponse(json.dumps(result_data))
    else:
        return HttpResponse("False")

def start_live_classroom(request):
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    session_years=SessionYearModel.object.all()
    return render(request,"staff_template/start_live_classroom.html",{"subjects":subjects,"session_years":session_years})

def start_live_classroom_process(request):
    session_year=request.POST.get("session_year")
    subject=request.POST.get("subject")

    subject_obj=Subjects.objects.get(id=subject)
    session_obj=SessionYearModel.object.get(id=session_year)
    checks=OnlineClassRoom.objects.filter(subject=subject_obj,session_years=session_obj,is_active=True).exists()
    if checks:
        data=OnlineClassRoom.objects.get(subject=subject_obj,session_years=session_obj,is_active=True)
        room_pwd=data.room_pwd
        roomname=data.room_name
    else:
        room_pwd=datetime.now().strftime('%Y%m-%d%H-%M%S-') + str(uuid4())
        roomname=datetime.now().strftime('%Y%m-%d%H-%M%S-') + str(uuid4())
        staff_obj=Staffs.objects.get(admin=request.user.id)
        onlineClass=OnlineClassRoom(room_name=roomname,room_pwd=room_pwd,subject=subject_obj,session_years=session_obj,started_by=staff_obj,is_active=True)
        onlineClass.save()

    return render(request,"staff_template/live_class_room_start.html",{"username":request.user.username,"password":room_pwd,"roomid":roomname,"subject":subject_obj.subject_name,"session_year":session_obj})


def returnHtmlWidget(request):
    return render(request,"widget.html")


def staff_subjects(request):
    staff = Staffs.objects.get(admin=request.user.id)
    subjects = Subjects.objects.filter(staff_id=staff.admin.id)

    subjects_data = []
    for subject in subjects:
        student_count = Students.objects.filter(course_id=subject.course_id).count()
        subjects_data.append({
            "subject_id": subject.id,  
            "subject_name": subject.subject_name,
            "course_name": subject.course_id.course_name,
            "student_count": student_count,
            "credit_hours": subject.credit_hours,
            "semester": subject.semester
        })

    return render(request, "staff_template/manage_my_subjects_template.html", {"subjects_data": subjects_data})


def view_students_in_subject(request, subject_id):
    try:
        subject = Subjects.objects.get(id=subject_id)
        students = Students.objects.filter(course_id=subject.course_id)

        return render(request, "staff_template/view_students_in_subject.html", {
            "subject": subject,
            "students": students
        })
    except Subjects.DoesNotExist:
        messages.error(request, "Subject not found")
        return HttpResponseRedirect(reverse("staff_subjects"))
    
def select_subjects(request):
    student = Students.objects.get(admin=request.user.id)
    available_subjects = Subjects.objects.filter(course_id=student.course_id)

    return render(request, "student_template/select_subjects.html", {
        "available_subjects": available_subjects
    })

def save_selected_subjects(request):
    if request.method != "POST":
        return HttpResponseRedirect(reverse("select_subjects"))
    else:
        selected_subjects = request.POST.getlist('subjects')
        student = Students.objects.get(admin=request.user.id)

        StudentSubjectSelection.objects.filter(student=student).delete()

        for subject_id in selected_subjects:
            subject = Subjects.objects.get(id=subject_id)
            StudentSubjectSelection.objects.create(student=student, subject=subject)

        messages.success(request, "Selected subjects submitted successfully. Waiting for advisor approval.")
        return HttpResponseRedirect(reverse("select_subjects"))


def advisor_pending_subjects(request):
    staff = Staffs.objects.get(admin=request.user.id)
    pending_requests = StudentSubjectSelection.objects.filter(
        student__advisor=staff, 
        is_approved__isnull=True
    )

    return render(request, "staff_template/advisor_pending_subjects.html", {
        "pending_requests": pending_requests
    })

def advisor_approve_subject(request, selection_id):
    selection = StudentSubjectSelection.objects.get(id=selection_id)
    selection.is_approved = True
    selection.save()

    NotificationStudent.objects.create(
        student_id=selection.student,
        message=f"Your subject '{selection.subject.subject_name}' has been approved by your advisor."
    )

    messages.success(request, "✅ Subject approved.")
    return redirect('advisor_pending_subjects')

def advisor_reject_subject(request, selection_id):
    selection = StudentSubjectSelection.objects.get(id=selection_id)
    selection.is_approved = False
    selection.save()

    # إشعار للطالب
    NotificationStudent.objects.create(
        student_id=selection.student,
        message=f"Your subject '{selection.subject.subject_name}' has been rejected by your advisor."
    )

    messages.success(request, "❌ Subject rejected.")
    return redirect('advisor_pending_subjects')

def manage_grade_components(request):
    staff = request.user.id
    subjects = Subjects.objects.filter(staff_id=staff) 
    components = GradeComponent.objects.filter(subject__staff_id=staff)

    if request.method == "POST":
        subject_id = request.POST.get('subject')
        name = request.POST.get('name')
        weight = float(request.POST.get('weight'))

        try:
            subject = Subjects.objects.get(id=subject_id)

            existing_components = GradeComponent.objects.filter(subject=subject)

            if name.strip().lower() == "final":
                if existing_components.filter(name__iexact="Final").exists():
                    messages.error(request, "❌ Final component already exists for this subject.")
                    return redirect('manage_grade_components')

            total_weight = sum(c.weight for c in existing_components) + weight
            if total_weight > 1.0:
                messages.error(request, "❌ Total weight exceeds 100%. Cannot add this component.")
                return redirect('manage_grade_components')

            GradeComponent.objects.create(subject=subject, name=name, weight=weight)
            messages.success(request, "✅ Successfully added grade component!")
            return redirect('manage_grade_components')

        except Exception as e:
            print(e)
            messages.error(request, "❌ Failed to add grade component.")

    return render(request, "staff_template/manage_grade_components.html", {
        "subjects": subjects,
        "components": components
    })



def add_student_result_components(request):
    staff = request.user.id
    subjects = Subjects.objects.filter(staff_id=staff)

    students = None
    components = None
    selected_subject = None

    if request.method == "POST":
        subject_id = request.POST.get('subject')
        selected_subject = Subjects.objects.get(id=subject_id)
        students = Students.objects.filter(course_id=selected_subject.course_id)
        components = GradeComponent.objects.filter(subject=selected_subject)

        if 'student' in request.POST:
            student_id = request.POST.get('student')
            total_score = 0

            try:
                student = Students.objects.get(admin=student_id)

                student_result, created = StudentResult.objects.get_or_create(
                    student_id=student,
                    subject_id=selected_subject,
                    defaults={'final_grade': 0, 'is_released': False}
                )

                StudentComponentResult.objects.filter(student_result=student_result).delete()

                for key in request.POST:
                    if key.startswith('component_'):
                        component_id = key.split('_')[1]
                        score = float(request.POST[key])

                        component = GradeComponent.objects.get(id=component_id)

                        weighted_score = (score * component.weight) / 100
                        total_score += weighted_score

                        StudentComponentResult.objects.create(
                            student_result=student_result,
                            grade_component=component,
                            score=score
                        )

                student_result.final_grade = total_score * 100
                student_result.is_released = False  
                student_result.save()

                messages.success(request, "Successfully added or updated result with component grades.")
                return redirect('add_student_result_components')

            except Exception as e:
                print(e)
                messages.error(request, "Failed to add or update student result.")
                return redirect('add_student_result_components')

    return render(request, "staff_template/add_student_result_components.html", {
        "subjects": subjects,
        "students": students,
        "components": components,
        "selected_subject": selected_subject,
    })

def calculate_grade_letter(final_grade):
    if final_grade == 0:
        return "GR"  
    if final_grade >= 90:
        return "A"
    elif final_grade >= 85:
        return "A-"
    elif final_grade >= 80:
        return "B+"
    elif final_grade >= 75:
        return "B"
    elif final_grade >= 70:
        return "B-"
    elif final_grade >= 65:
        return "C+"
    elif final_grade >= 60:
        return "C"
    elif final_grade >= 55:
        return "D+"
    elif final_grade >= 50:
        return "D"
    else:
        return "F"  

def manage_release_results(request):
    staff = request.user.id
    subjects = Subjects.objects.filter(staff_id=staff)
    results = StudentResult.objects.filter(subject_id__in=subjects, is_released=False)

    return render(request, "staff_template/manage_release_results.html", {
        "results": results
    })

def release_result(request, result_id):
    result = get_object_or_404(StudentResult, id=result_id)

    result.grade_letter = calculate_grade_letter(result.final_grade)
    result.is_released = True
    result.save()

    messages.success(request, "Result released successfully with grade letter assigned.")
    return HttpResponseRedirect(reverse('manage_release_results'))

def manage_subject_final_results(request):
    staff = request.user.id
    subjects = Subjects.objects.filter(staff_id=staff)
    return render(request, "staff_template/manage_subject_final_results.html", {
        "subjects": subjects
    })

def view_final_results_for_subject(request, subject_id):
    subject = get_object_or_404(Subjects, id=subject_id)
    results = StudentResult.objects.filter(subject_id=subject)
    components = GradeComponent.objects.filter(subject=subject)

    for result in results:
        result.temp_grade_letter = calculate_grade_letter(result.final_grade)

    return render(request, "staff_template/view_final_results_for_subject.html", {
        "subject": subject,
        "results": results,
        "components": components,
    })


def release_final_results_for_subject(request, subject_id):
    subject = get_object_or_404(Subjects, id=subject_id)
    results = StudentResult.objects.filter(subject_id=subject, is_released=False)

    for result in results:
        result.grade_letter = calculate_grade_letter(result.final_grade)
        result.is_released = True
        result.save()

    messages.success(request, f"All results for {subject.subject_name} have been released.")
    return HttpResponseRedirect(reverse('manage_subject_final_results'))

def edit_student_component_results(request, student_result_id):
    student_result = get_object_or_404(StudentResult, id=student_result_id)
    components = StudentComponentResult.objects.filter(student_result=student_result)

    if request.method == "POST":
        total_score = 0

        for component_result in components:
            new_score = float(request.POST.get(f'component_{component_result.id}', 0))
            component_result.score = new_score
            component_result.save()

            weighted_score = (new_score * component_result.grade_component.weight) / 100
            total_score += weighted_score

        student_result.final_grade = total_score * 100
        student_result.is_released = False  
        student_result.save()

        messages.success(request, "Successfully updated component scores and recalculated final grade.")
        return HttpResponseRedirect(reverse('manage_subject_final_results'))

    return render(request, "staff_template/edit_student_component_results.html", {
        "student_result": student_result,
        "components": components
    })


def delete_grade_component(request, component_id):
    component = get_object_or_404(GradeComponent, id=component_id)

    try:
        component.delete()
        messages.success(request, "✅ Successfully deleted grade component.")
    except:
        messages.error(request, "❌ Failed to delete grade component.")

    return redirect('manage_grade_components')


from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import HttpResponse

def export_final_results_pdf(request, subject_id):
    subject = get_object_or_404(Subjects, id=subject_id)
    results = StudentResult.objects.filter(subject_id=subject)
    components = GradeComponent.objects.filter(subject=subject)

    template_path = 'staff_template/pdf_template.html'
    context = {
        'subject': subject,
        'results': results,
        'components': components,
    }

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Final_Results_{subject.subject_name}.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def export_final_results_excel(request, subject_id):
    subject = get_object_or_404(Subjects, id=subject_id)
    results = StudentResult.objects.filter(subject_id=subject)
    components = GradeComponent.objects.filter(subject=subject)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"{subject.subject_name} Results"

    # Create header
    headers = ['Student ID', 'Student Name']
    headers += [component.name for component in components]
    headers += ['Final Grade', 'Grade Letter', 'Status']

    worksheet.append(headers)

    # Styling the header
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')

    # Fill data
    for result in results:
        row = [
            result.student_id.admin.id,
            f"{result.student_id.admin.first_name} {result.student_id.admin.last_name}"
        ]

        for component in components:
            try:
                score = StudentComponentResult.objects.get(
                    student_result=result,
                    grade_component=component
                ).score
            except StudentComponentResult.DoesNotExist:
                score = 'N/A'
            row.append(score)

        row.append(result.final_grade)
        row.append(result.grade_letter if result.grade_letter else '-')
        row.append('Released' if result.is_released else 'Pending')

        worksheet.append(row)

    # Auto width adjustment
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(column)].width = max_length + 3

    # Add AutoFilter
    worksheet.auto_filter.ref = worksheet.dimensions

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Final_Results_{subject.subject_name}.xlsx"'
    workbook.save(response)
    return response


from django.shortcuts import render, redirect
from django.contrib import messages
from obs_system_app.models import Subjects, SubjectResource
from django.contrib.auth.decorators import login_required

@login_required
def upload_subject_resource(request):
    staff = request.user.staffs
    subjects = Subjects.objects.filter(staff_id=staff.admin.id)

    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        title = request.POST.get('title')
        file = request.FILES.get('file')

        try:
            subject = Subjects.objects.get(id=subject_id)
            SubjectResource.objects.create(subject=subject, title=title, file=file)
            messages.success(request, 'Resource uploaded successfully.')
            return redirect('upload_subject_resource')
        except Exception as e:
            messages.error(request, f'Upload failed: {e}')

    return render(request, 'staff_template/upload_subject_resource.html', {'subjects': subjects})

def assign_advisor_to_students(request):
    staff = Staffs.objects.get(admin=request.user.id)
    students_without_advisor = Students.objects.filter(advisor__isnull=True)
    return render(request, "staff_template/assign_advisor.html", {
        "students": students_without_advisor
    })

def confirm_assign_advisor(request, student_id):
    try:
        student = Students.objects.get(id=student_id)
        staff = Staffs.objects.get(admin=request.user.id)
        student.advisor = staff
        student.save()
        messages.success(request, f"You have been assigned as an advisor to {student.admin.get_full_name()}.")
    except Exception as e: 
        messages.error(request, "Failed to assign advisor.")
    return redirect("assign_advisor_to_students")

