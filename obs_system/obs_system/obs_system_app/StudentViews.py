import datetime

from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt

from obs_system_app.models import Students, Courses, Subjects, CustomUser, Attendance, AttendanceReport, \
    LeaveReportStudent, FeedBackStudent, NotificationStudent, OnlineClassRoom, SessionYearModel, StudentSubjectSelection,\
    GradeComponent, StudentResult, StudentComponentResult,DownloadPeriod,SubjectResource,SubjectRegistrationPeriod


def student_home(request):
    student_obj = Students.objects.get(admin=request.user.id)

    update_student_year_semester(student_obj)

    attendance_total = AttendanceReport.objects.filter(student_id=student_obj).count()
    attendance_present = AttendanceReport.objects.filter(student_id=student_obj, status=True).count()
    attendance_absent = AttendanceReport.objects.filter(student_id=student_obj, status=False).count()
    course = Courses.objects.get(id=student_obj.course_id.id)
    subjects = Subjects.objects.filter(course_id=course).count()
    subjects_data = Subjects.objects.filter(course_id=course)
    session_obj = SessionYearModel.object.get(id=student_obj.session_year_id.id)
    class_room = OnlineClassRoom.objects.filter(subject__in=subjects_data, is_active=True, session_years=session_obj)

    subject_name = []
    data_present = []
    data_absent = []
    subject_data = Subjects.objects.filter(course_id=student_obj.course_id)
    for subject in subject_data:
        attendance = Attendance.objects.filter(subject_id=subject.id)
        attendance_present_count = AttendanceReport.objects.filter(attendance_id__in=attendance, status=True, student_id=student_obj.id).count()
        attendance_absent_count = AttendanceReport.objects.filter(attendance_id__in=attendance, status=False, student_id=student_obj.id).count()
        subject_name.append(subject.subject_name)
        data_present.append(attendance_present_count)
        data_absent.append(attendance_absent_count)

    return render(request, "student_template/student_home_template.html", {
        "total_attendance": attendance_total,
        "attendance_absent": attendance_absent,
        "attendance_present": attendance_present,
        "subjects": subjects,
        "data_name": subject_name,
        "data1": data_present,
        "data2": data_absent,
        "class_room": class_room,
        "current_year": student_obj.current_year,
        "current_semester": student_obj.current_semester,
        "advisor": student_obj.advisor,
    })
def join_class_room(request,subject_id,session_year_id):
    session_year_obj=SessionYearModel.object.get(id=session_year_id)
    subjects=Subjects.objects.filter(id=subject_id)
    if subjects.exists():
        session=SessionYearModel.object.filter(id=session_year_obj.id)
        if session.exists():
            subject_obj=Subjects.objects.get(id=subject_id)
            course=Courses.objects.get(id=subject_obj.course_id.id)
            check_course=Students.objects.filter(admin=request.user.id,course_id=course.id)
            if check_course.exists():
                session_check=Students.objects.filter(admin=request.user.id,session_year_id=session_year_obj.id)
                if session_check.exists():
                    onlineclass=OnlineClassRoom.objects.get(session_years=session_year_id,subject=subject_id)
                    return render(request,"student_template/join_class_room_start.html",{"username":request.user.username,"password":onlineclass.room_pwd,"roomid":onlineclass.room_name})

                else:
                    return HttpResponse("This Online Session is Not For You")
            else:
                return HttpResponse("This Subject is Not For You")
        else:
            return HttpResponse("Session Year Not Found")
    else:
        return HttpResponse("Subject Not Found")


def student_view_attendance(request):
    student=Students.objects.get(admin=request.user.id)
    course=student.course_id
    subjects=Subjects.objects.filter(course_id=course)
    return render(request,"student_template/student_view_attendance.html",{"subjects":subjects})

def student_view_attendance_post(request):
    subject_id=request.POST.get("subject")
    start_date=request.POST.get("start_date")
    end_date=request.POST.get("end_date")

    start_data_parse=datetime.datetime.strptime(start_date,"%Y-%m-%d").date()
    end_data_parse=datetime.datetime.strptime(end_date,"%Y-%m-%d").date()
    subject_obj=Subjects.objects.get(id=subject_id)
    user_object=CustomUser.objects.get(id=request.user.id)
    stud_obj=Students.objects.get(admin=user_object)

    attendance=Attendance.objects.filter(attendance_date__range=(start_data_parse,end_data_parse),subject_id=subject_obj)
    attendance_reports=AttendanceReport.objects.filter(attendance_id__in=attendance,student_id=stud_obj)
    return render(request,"student_template/student_attendance_data.html",{"attendance_reports":attendance_reports})

def student_apply_leave(request):
    staff_obj = Students.objects.get(admin=request.user.id)
    leave_data=LeaveReportStudent.objects.filter(student_id=staff_obj)
    return render(request,"student_template/student_apply_leave.html",{"leave_data":leave_data})

def student_apply_leave_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("student_apply_leave"))
    else:
        leave_date=request.POST.get("leave_date")
        leave_msg=request.POST.get("leave_msg")

        student_obj=Students.objects.get(admin=request.user.id)
        try:
            leave_report=LeaveReportStudent(student_id=student_obj,leave_date=leave_date,leave_message=leave_msg,leave_status=0)
            leave_report.save()
            messages.success(request, "Successfully Applied for Leave")
            return HttpResponseRedirect(reverse("student_apply_leave"))
        except:
            messages.error(request, "Failed To Apply for Leave")
            return HttpResponseRedirect(reverse("student_apply_leave"))


def student_feedback(request):
    staff_id=Students.objects.get(admin=request.user.id)
    feedback_data=FeedBackStudent.objects.filter(student_id=staff_id)
    return render(request,"student_template/student_feedback.html",{"feedback_data":feedback_data})

def student_feedback_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("student_feedback"))
    else:
        feedback_msg=request.POST.get("feedback_msg")

        student_obj=Students.objects.get(admin=request.user.id)
        try:
            feedback=FeedBackStudent(student_id=student_obj,feedback=feedback_msg,feedback_reply="")
            feedback.save()
            messages.success(request, "Successfully Sent Feedback")
            return HttpResponseRedirect(reverse("student_feedback"))
        except:
            messages.error(request, "Failed To Send Feedback")
            return HttpResponseRedirect(reverse("student_feedback"))

import datetime

def calculate_student_year_and_semester(student):
    today = datetime.date.today()

    if not student.enrollment_date:
        return "N/A", "N/A"

    enrollment_date = student.enrollment_date
    years_difference = today.year - enrollment_date.year

    if today.month < 9:
        years_difference -= 1

    student_year = max(1, years_difference + 1)

    if today.month >= 9 or today.month <= 2:
        semester = 1  # Fall Semester
    else:
        semester = 2  # Spring Semester

    return student_year, semester

def student_profile(request):
    user = CustomUser.objects.get(id=request.user.id)
    student = Students.objects.get(admin=user)

    student_year, student_semester = calculate_student_year_and_semester(student)

    return render(request, "student_template/student_profile.html", {
        "user": user,
        "student": student,
        "student_year": student_year,
        "student_semester": student_semester
    })

def student_profile_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("student_profile"))
    else:
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        password=request.POST.get("password")
        address=request.POST.get("address")
        try:
            customuser=CustomUser.objects.get(id=request.user.id)
            customuser.first_name=first_name
            customuser.last_name=last_name
            if password!=None and password!="":
                customuser.set_password(password)
            customuser.save()

            student=Students.objects.get(admin=customuser)
            student.address=address
            student.save()
            messages.success(request, "Successfully Updated Profile")
            return HttpResponseRedirect(reverse("student_profile"))
        except:
            messages.error(request, "Failed to Update Profile")
            return HttpResponseRedirect(reverse("student_profile"))

@csrf_exempt
def student_fcmtoken_save(request):
    token=request.POST.get("token")
    try:
        student=Students.objects.get(admin=request.user.id)
        student.fcm_token=token
        student.save()
        return HttpResponse("True")
    except:
        return HttpResponse("False")

def student_all_notification(request):
    student=Students.objects.get(admin=request.user.id)
    notifications=NotificationStudent.objects.filter(student_id=student.id)
    return render(request,"student_template/all_notification.html",{"notifications":notifications})


def select_subjects(request):
    import datetime
    from obs_system_app.recommendation.utils import load_similarity_matrix, recommend

    student = Students.objects.get(admin=request.user.id)
    today = datetime.date.today()

    # Calculate current student year and semester
    student_year, student_semester = calculate_student_year_and_semester(student)
    semester_number = (student_year - 1) * 2 + student_semester

    # Check if registration period is open
    registration_open = SubjectRegistrationPeriod.objects.filter(
        session_year=student.session_year_id,
        semester=student_semester,
        start_date__lte=today,
        end_date__gte=today,
        is_active=True
    ).exists()

    # Subjects for the current semester
    compulsory_subjects = Subjects.objects.filter(
        course_id=student.course_id,
        semester=semester_number,
        subject_type="Compulsory"
    )

    elective_subjects = Subjects.objects.filter(
        course_id=student.course_id,
        semester=semester_number,
        subject_type="Elective"
    )

    # Max electives per semester
    elective_limits = {
        5: 1,
        6: 2,
        7: 4,
        8: 4
    }
    max_electives = elective_limits.get(semester_number, 0)

    # Subjects previously selected by the student (approved or not)
    selected_subjects = StudentSubjectSelection.objects.filter(student=student).values_list('subject_id', flat=True)
    approved_subjects = StudentSubjectSelection.objects.filter(student=student)

    #  Recommendations using item-based CF
    try:
        similarity_df = load_similarity_matrix()
        valid_ids = set(similarity_df.index)

        # Subjects taken by the student used for recommendation
        filtered_subjects_taken = [sid for sid in selected_subjects if sid in valid_ids]

        # Cold Start: If number of subjects < 2, add dummy records
        if len(filtered_subjects_taken) < 2:
            pseudo_history = list(valid_ids)[:2]
            filtered_subjects_taken += pseudo_history
            messages.info(request, "🔍 Recommendations are based on limited history. Add more electives to improve accuracy.")

        # Generate recommendations
        recommended_ids = recommend(filtered_subjects_taken, similarity_df, top_n=5)

        # Filter recommendations to be from the same semester only
        recommended_subjects = Subjects.objects.filter(
            id__in=recommended_ids,
            subject_type="Elective",
            semester=semester_number
        )

    except Exception as e:
        print(f"Recommendation failed: {e}")
        recommended_subjects = []

    # Calculate GPA
    cumulative_gpa = calculate_gpa(student)

    return render(request, "student_template/select_subjects.html", {
        "compulsory_subjects": compulsory_subjects,
        "elective_subjects": elective_subjects,
        "selected_subjects": selected_subjects,
        "approved_subjects": approved_subjects,
        "student_year": student_year,
        "student_semester": student_semester,
        "user": request.user,
        "student": student,
        "cumulative_gpa": cumulative_gpa,
        "registration_open": registration_open,
        "advisor": student.advisor,
        "max_electives": max_electives,
        "recommended_subjects": recommended_subjects,
    })




from django.contrib import messages
from django.http import HttpResponseRedirect
from django.urls import reverse

def save_selected_subjects(request):
    if request.method != "POST":
        return HttpResponseRedirect(reverse("select_subjects"))
    else:
        student = Students.objects.get(admin=request.user.id)

        today = datetime.date.today()
        registration_open = SubjectRegistrationPeriod.objects.filter(
            session_year=student.session_year_id,
            semester=student.current_semester,
            start_date__lte=today,
            end_date__gte=today,
            is_active=True
        ).exists()

        if not registration_open:
            messages.error(request, "❌ Subject registration period is closed. You cannot submit now.")
            return HttpResponseRedirect(reverse("select_subjects"))

        selected_subjects = request.POST.getlist('subjects')

        # Check that pre-approved subjects are not being re-submitted
        approved_subjects = StudentSubjectSelection.objects.filter(student=student, is_approved=True).values_list('subject_id', flat=True)
        for subj_id in selected_subjects:
            if int(subj_id) in approved_subjects:
                messages.error(request, "❌ One or more of the selected subjects have already been approved. You cannot modify them.")
                return HttpResponseRedirect(reverse("select_subjects"))

        # Delete and re-submit requests (only unapproved ones)
        StudentSubjectSelection.objects.filter(student=student, is_approved__isnull=True).delete()

        for subject_id in selected_subjects:
            subject = Subjects.objects.get(id=subject_id)
            StudentSubjectSelection.objects.create(student=student, subject=subject)

        messages.success(request, "✅ Your subjects have been submitted for advisor approval.")
        return HttpResponseRedirect(reverse("select_subjects"))

def student_view_results(request):
    student = Students.objects.get(admin=request.user.id)
    all_results = StudentResult.objects.filter(student_id=student)
    components = GradeComponent.objects.all()

    results_with_components = []
    used_subjects = set()

    for result in all_results:
        if result.subject_id.id not in used_subjects:
            components_scores = StudentComponentResult.objects.filter(student_result=result)
            results_with_components.append({
                "result": result,
                "components": components_scores
            })
            used_subjects.add(result.subject_id.id)

    gpa = calculate_gpa(student)

    released_results = StudentResult.objects.filter(student_id=student, is_released=True).distinct('subject_id')
    total_credits = sum([r.subject_id.credit_hours for r in released_results])

    return render(request, "student_template/student_view_results.html", {
        "results_with_components": results_with_components,
        "components": components,
        "gpa": gpa,
        "total_credits": total_credits,
    })


from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import HttpResponse

def student_export_results_pdf(request):
    student = Students.objects.get(admin=request.user.id)
    results = StudentResult.objects.filter(student_id=student).distinct('subject_id')
    components = GradeComponent.objects.all()

    gpa = calculate_gpa(student)

    template_path = 'student_template/student_pdf_template.html'
    context = {
        'student': student,
        'results': results,
        'components': components,
        'gpa': gpa,
    }

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{student.admin.username}_Transcript.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response



import openpyxl
from openpyxl.utils import get_column_letter

def student_export_results_excel(request):
    student = Students.objects.get(admin=request.user.id)
    results = StudentResult.objects.filter(student_id=student).distinct('subject_id')
    components = GradeComponent.objects.all()

    gpa = calculate_gpa(student)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "My Results"

    # Create header
    headers = ['Course Code', 'Course Name']
    headers += [component.name for component in components]
    headers += ['Final Grade', 'Grade Letter', 'Status']

    worksheet.append(headers)

    # Fill data
    for result in results:
        row = [
            f"{result.subject_id.subject_code_dep}{result.subject_id.id}",
            result.subject_id.subject_name
        ]

        for component in components:
            try:
                score = StudentComponentResult.objects.get(
                    student_result=result,
                    grade_component=component
                ).score
                row.append(round(score, 2))
            except StudentComponentResult.DoesNotExist:
                row.append('N/A')

        row.append(round(result.final_grade, 2))
        row.append(result.grade_letter if result.grade_letter else '-')
        row.append('Released' if result.is_released else 'Pending')

        worksheet.append(row)

    # Insert GPA
    worksheet.append([])
    worksheet.append(["", "", "", "", "", "", f"GPA: {gpa}"])

    # Auto width adjustment
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(column)].width = max_length + 3

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{student.admin.username}_Transcript.xlsx"'
    workbook.save(response)
    return response



def calculate_grade_point(final_grade):
    if final_grade >= 90:
        return 4.0
    elif final_grade >= 85:
        return 3.7
    elif final_grade >= 80:
        return 3.3
    elif final_grade >= 75:
        return 3.0
    elif final_grade >= 70:
        return 2.7
    elif final_grade >= 65:
        return 2.3
    elif final_grade >= 60:
        return 2.0
    elif final_grade >= 55:
        return 1.7
    elif final_grade >= 50:
        return 1.0
    else:
        return 0.0

def calculate_gpa(student):
    results = StudentResult.objects.filter(student_id=student, is_released=True).distinct('subject_id')

    total_points = 0
    total_credits = 0

    for result in results:
        if result.final_grade is not None:
            grade_point = calculate_grade_point(result.final_grade)
            credit_hours = result.subject_id.credit_hours
            total_points += grade_point * credit_hours
            total_credits += credit_hours

    if total_credits == 0:
        return 0.0

    gpa = total_points / total_credits
    return round(gpa, 2)

def student_view_gpa_detailed(request):
    student = Students.objects.get(admin=request.user.id)

    cumulative_gpa = calculate_gpa(student)  # Cumulative GPA

    # Get distinct semesters based on subjects
    semesters = Subjects.objects.values_list('semester', flat=True).distinct().order_by('semester')

    semester_gpas = []

    for sem in semesters:
        # Get subjects in this semester
        subjects_in_semester = Subjects.objects.filter(semester=sem)

        # Get the student's results in subjects of this semester that have been released
        results = StudentResult.objects.filter(
            student_id=student,
            subject_id__in=subjects_in_semester,
            is_released=True
        )

        course_count = results.count()
        total_credits = 0
        total_ects = 0.0
        total_points = 0.0

        for result in results:
            credit_hours = result.subject_id.credit_hours
            ects_credits = result.subject_id.ects_credits
            grade_point = calculate_grade_point(result.final_grade)

            total_credits += credit_hours
            total_ects += ects_credits
            total_points += grade_point * credit_hours

        semester_gpa = 0.0
        if total_credits > 0:
            semester_gpa = total_points / total_credits

        semester_name = f"Semester {sem}"

        semester_gpas.append({
            'semester': sem,
            'semester_name': semester_name,
            'course_count': course_count,
            'total_credits': total_credits,
            'total_ects': round(total_ects, 2),
            'gpa': round(semester_gpa, 2)
        })

    context = {
        'cumulative_gpa': round(cumulative_gpa, 2),
        'semester_gpas': semester_gpas,
        'student': student
    }

    return render(request, 'student_template/student_gpa.html', context)

def update_student_year_semester(student):
    today = datetime.date.today()
    enrollment_date = student.enrollment_date
    if not enrollment_date:
        return  

    years_difference = today.year - enrollment_date.year
    if today.month < 9:
        years_difference -= 1
    current_year = max(1, years_difference + 1)

    if today.month >= 9 or today.month <= 2:
        current_semester = 1
    else:
        current_semester = 2

    student.current_year = current_year
    student.current_semester = current_semester
    student.save()


def student_subject_resources(request):
    student = Students.objects.get(admin=request.user.id)
    subjects = Subjects.objects.filter(course_id=student.course_id, semester=student.current_semester)
    resources = SubjectResource.objects.filter(subject__in=subjects)

    return render(request, 'student_template/subject_resources.html', {
        'resources': resources
    })